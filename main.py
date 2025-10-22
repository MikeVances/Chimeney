# main.py
# Flask-сервер для подбора вентиляционной шахты
# Контракт: POST /api/select -> {results: [{article, name, quantity}]}

from flask import Flask, request, jsonify, send_from_directory, send_file
import json
import os
import re
from typing import Dict, Any, List

from io import BytesIO
from openpyxl import Workbook

app = Flask(__name__)

# === Загрузка каталога ===
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
CATALOG_PATH = os.path.join(BASE_DIR, "data", "komplektuyushchie.json")

with open(CATALOG_PATH, "r", encoding="utf-8") as f:
    raw = json.load(f)

if isinstance(raw, dict):
    code_mapping: Dict[str, str] = raw.get("_code_mapping", {})
    items: List[Dict[str, Any]] = raw.get("items")
    if not items:
        # fallback: если данные пришли «плоским» массивом внутри словаря
        items = [v for v in raw.values() if isinstance(v, dict)]
else:
    # fallback: если это список словарей без служебного блока
    items = raw
    code_mapping = {}

# Индексы
by_art: Dict[str, Dict[str, Any]] = {}
for it in items:
    art = str(it.get("artikul") or it.get("article") or "").strip()
    if art:
        by_art[art] = it


# === Утилиты ===

def _norm(s: Any) -> str:
    if s is None:
        return ""
    s = str(s).strip().lower()
    # без агрессивной замены букв — фронт шлёт нормализованные ключи
    return s


def _phase_from_motor(motor: str) -> str:
    # тип_мотора: '6e' -> 1, '6d' -> 3
    motor = _norm(motor)
    return {"6e": "1", "6d": "3"}.get(motor, "")


def _find(predicate):
    for it in items:
        try:
            if predicate(it):
                return it
        except Exception:
            continue
    return None


def _find_all(predicate):
    out = []
    for it in items:
        try:
            if predicate(it):
                out.append(it)
        except Exception:
            continue
    return out


def _name_has(it: Dict[str, Any], *need):
    name = _norm(it.get("name"))
    return all(w in name for w in need)


def pick_membrana_by_diam(diam: str):
    """Ищем мембрану строго под диаметр: сначала по полю diameter, затем по имени."""
    # по атрибуту diameter
    it = _find(lambda x: _norm(x.get("category")) == "мембрана" and str(x.get("diameter", "")).strip() == diam)
    if it:
        return it
    # по имени (например, "VB-710" или просто "710")
    return _find(lambda x: _norm(x.get("category")) == "мембрана" and (diam in _norm(x.get("name"))))


def pick_lenta():
    """Ищем универсальную битумную ленту по имени."""
    return _find(lambda x: "лента" in _norm(x.get("name")))


def pick_hermetic(membrana: bool, lenta: bool):
    out = []
    if membrana:
        it = _find(lambda x: _norm(x.get("category")) == "мембрана" or "мембран" in _norm(x.get("name")))
        if it:
            out.append(it)
    if lenta:
        it = _find(lambda x: "лента" in _norm(x.get("name")))  # иногда категория другая
        if it:
            out.append(it)
    return out


# === Формирование ключа для _code_mapping ===

def build_code_key(payload: Dict[str, Any]) -> str:
    tip = payload.get("tip")  # VBV/VBP/VBA/VBR
    diam = str(payload.get("diametr", "")).strip()
    klapan = payload.get("tip_klapana")  # pov/grav/dvustv
    rasp = payload.get("raspolozhenie")  # niz/verh (только для pov)
    grav_variant = payload.get("grav_variant")  # vnut/vnesh (только для grav)
    motor = payload.get("tip_motora")  # 6e/6d
    power = str(payload.get("moshchnost", "")).strip()  # 370/750

    if not tip or not diam:
        return ""

    tip = tip.upper()  # VBV/VBP/VBA/VBR

    # В большинстве ключей требуется мощность+фазность (для активных/вытяжных)
    phase = _phase_from_motor(motor) if motor else ""

    # Ветка по типу клапана
    if klapan == "pov":
        if not power or not phase or not rasp:
            return ""
        pos = {"niz": "niz", "verh": "verh"}.get(rasp)
        if not pos:
            return ""
        return f"{tip}_{diam}_{power}_{phase}_pov_{pos}"

    if klapan == "grav":
        if not power or not phase or not grav_variant:
            return ""
        gv = {"vnut": "grav_vnut", "vnesh": "grav_vnesh"}.get(grav_variant)
        if not gv:
            return ""
        return f"{tip}_{diam}_{power}_{phase}_{gv}"

    if klapan == "dvustv":
        # для двустворчатого клапана в ряде каталогов ключ может не требовать power/phase
        # оставляем консервативный вариант без них, при необходимости расширим
        return f"{tip}_{diam}_dvustv"

    return ""


# === Подбор доп.комплектующих ===

def pick_drive_for_pov(diam: str):
    # Заглушка: на текущей итерации приводы не подбираем вовсе
    return None


def pick_top_part(kind: str, diam: str):
    # kind: 'zont'|'rastrub' (универсальные, без привязки к диаметру)
    kind = _norm(kind)
    if kind == "zont":
        # сначала по категории, затем по имени
        if "89174" in by_art:
            return by_art["89174"]
        it = _find(lambda x: _name_has(x, "комплект", "зонт"))
        if it:
            return it    
        it = _find(lambda x: _norm(x.get("category")) == "зонт")
        if it:
            return it
        return _find(lambda x: _name_has(x, "зонт"))
    if kind == "rastrub":
        # Раструб привязан к диаметру: сначала ищем по атрибутам, затем по имени
        # 1) по category/diameter, если в каталоге заполнены поля
        it = _find(lambda x: _norm(x.get("category")) == "раструб" and str(x.get("diameter", "")).strip() == diam)
        if it:
            return it
        # 2) по имени + диаметр (часто лежит в "прочее")
        return _find(lambda x: _name_has(x, "раструб") and (diam in _norm(x.get("name"))))
    return None


def pick_udlinenie_sections(diam: str, meters: int):
    if not meters:
        return None, 0
    # Секции 1 м: category=='секция', type=='VB', diameter==diam
    it = _find(lambda x: _norm(x.get("category")) == "секция" and _norm(x.get("type")) == "vb" and str(x.get("diameter", "")).strip() == diam)
    if it:
        return it, int(meters)
    # fallback по имени
    it = _find(lambda x: _name_has(x, "секция") and ("vb" in _norm(x.get("type")) or " vb" in _norm(x.get("name"))) and diam in _norm(x.get("name")))
    return it, int(meters) if it else (None, 0)


def pick_hermetic(membrana: bool, lenta: bool):
    out = []
    if membrana:
        it = _find(lambda x: _norm(x.get("category")) == "мембрана" or "мембран" in _norm(x.get("name")))
        if it:
            out.append(it)
    if lenta:
        it = _find(lambda x: "лента" in _norm(x.get("name")))  # иногда категория другая
        if it:
            out.append(it)
    return out


def pick_avtomat(target_amps: float):
    """Выбираем автомат по номиналу тока: поддерживаются одиночные значения и диапазоны (например, 1.0-1.6A).
    Парсим ток из имени (форматы: "3 A", "3A", "2,4 A", т.п.).
    """
    candidates_exact = []   # (val, item)
    candidates_upper = []   # (upper_bound, item) for ranges and singles used as fallback

    for x in items:
        if not _name_has(x, "автомат"):
            continue
        name = x.get("name", "")
        # 1) Диапазон, например: "1.0-1.6А" (пробелы и разные дефисы допустимы)
        m_range = re.search(r"(\d+(?:[\.,]\d+)?)\s*[-–—]\s*(\d+(?:[\.,]\d+)?)(?:\s*A|A)?", name, flags=re.IGNORECASE)
        if m_range:
            try:
                v1 = float(m_range.group(1).replace(',', '.'))
                v2 = float(m_range.group(2).replace(',', '.'))
                lo, hi = (v1, v2) if v1 <= v2 else (v2, v1)
            except Exception:
                lo = hi = None
            if lo is not None and hi is not None:
                # если искомый ток внутри диапазона — это идеальное совпадение
                if lo - 1e-9 <= target_amps <= hi + 1e-9:
                    return x
                # иначе используем верхнюю границу как кандидат для "минимального большего"
                candidates_upper.append((hi, x))
                continue
        # 2) Одиночное значение: "3 A" или "3A"
        m_single = re.search(r"(\d+(?:[\.,]\d+)?)\s*A", name, flags=re.IGNORECASE) or \
                   re.search(r"(\d+(?:[\.,]\d+)?)A", name, flags=re.IGNORECASE)
        if m_single:
            try:
                val = float(m_single.group(1).replace(',', '.'))
            except Exception:
                continue
            # сначала попытаемся найти точное совпадение
            candidates_exact.append((val, x))
            candidates_upper.append((val, x))

    # точное совпадение среди одиночных
    for val, x in candidates_exact:
        if abs(val - target_amps) < 1e-6:
            return x

    # минимальный больший среди всех верхних границ (и одиночных значений)
    greater = [(val, x) for val, x in candidates_upper if val >= target_amps - 1e-9]
    if greater:
        val, x = sorted(greater, key=lambda t: t[0])[0]
        return x

    return None


def pick_kapleu(diam: str):
    # Универсальный подбор
    return _find(lambda x: _name_has(x, "каплеулав"))


def pick_korona():
    return _find(lambda x: _name_has(x, "корона"))


# New helper for listing available drives
def list_available_drives(max_items: int = 50):
    """Возвращает список приводов для подсказки пользователю.
    1) Сначала берём артикулы из белого списка (если есть в каталоге);
    2) Затем добавляем найденные по строгому фильтру ("электропривод"|категория "привод"),
       исключая любые секции/шахты/клапаны.
    """
    # Белый список (из номенклатуры):
    whitelist = {
        "7547", "29299", "3557", "21370", "1191",
        "84015", "84935", "84016", "84934",
    }

    picked = []
    seen = set()

    # 1) по whitelists — в фиксированном порядке
    for art in [
        "7547", "29299", "3557", "21370", "1191",
        "84015", "84935", "84016", "84934",
    ]:
        it = by_art.get(art)
        if it and art not in seen:
            picked.append(it)
            seen.add(art)
        if len(picked) >= max_items:
            return picked

    # 2) строгий фильтр по каталогу
    def is_true_drive(x: Dict[str, Any]) -> bool:
        nm = _norm(x.get("name"))
        cat = _norm(x.get("category"))
        # ключевые слова для приводов
        is_drive_kw = ("электропривод" in nm) or ("привод" in cat) or ("bvm" in nm)
        if not is_drive_kw:
            return False
        # исключаем явные не-приводы
        bad = ("секция" in nm) or ("vbv" in nm) or ("vba" in nm) or ("vbr" in nm) or ("клапан" in nm)
        return not bad

    for x in items:
        art = str(x.get("artikul") or x.get("article") or "").strip()
        if art and art not in seen and is_true_drive(x):
            picked.append(x)
            seen.add(art)
        if len(picked) >= max_items:
            break

    return picked[:max_items]

# Helper for VBR подмешивание
def pick_vbr_podmesh(diam: str):
    # Ищем секцию подмешивания по имени + диаметр
    return _find(lambda x: _name_has(x, "подмешив") and (diam in _norm(x.get("name"))))


# --- New helpers and refactored logic for export/select ---

def _price_of(art: str) -> float:
    it = by_art.get(str(art).strip())
    if not it:
        return 0.0
    for key in ("price", "цена", "cost"):
        v = it.get(key)
        if v is None:
            continue
        try:
            return float(str(v).replace(' ', '').replace(',', '.'))
        except Exception:
            continue
    return 0.0


def _select_components(payload_in: Dict[str, Any]):
    payload = dict(payload_in or {})
    messages: List[str] = []

    tip = payload.get("tip")
    diam = str(payload.get("diametr", "")).strip()
    klapan = payload.get("tip_klapana")

    # Бэкенд-ограничения: приведение параметров к допустимым
    if tip == "VBA" and klapan != "pov":
        payload["tip_klapana"] = "pov"
        klapan = "pov"
        messages.append("Для VBA выбран поворотный клапан (гравитационного не бывает)")

    if tip == "VBR":
        if klapan != "pov":
            payload["tip_klapana"] = "pov"
            klapan = "pov"
            messages.append("Для VBR выбран поворотный клапан (гравитационного/двустворчатого не бывает)")
        if _norm(payload.get("raspolozhenie")) != "niz":
            payload["raspolozhenie"] = "niz"
            messages.append("Для VBR расположение клапана всегда 'низ'")

    if tip == "VBP":
        if klapan != "pov":
            payload["tip_klapana"] = "pov"
            klapan = "pov"
            messages.append("Для VBP доступен только поворотный клапан")
        if _norm(payload.get("raspolozhenie")) != "niz":
            payload["raspolozhenie"] = "niz"
            messages.append("Для VBP расположение клапана всегда 'низ'")

    required_power = {"560": "370", "710": "370", "800": "750"}.get(diam)

    if not tip or not diam or not klapan:
        return [], ["Не хватает обязательных параметров"]

    needs_motor = (tip in ("VBV", "VBA", "VBR")) and (klapan in ("pov", "grav"))
    if needs_motor:
        if not payload.get("tip_motora"):
            return [], ["Не хватает параметров: укажите тип мотора (6E/6D)"]
        if required_power:
            cur_power = str(payload.get("moshchnost", "")).strip()
            if not cur_power:
                payload["moshchnost"] = required_power
                messages.append(f"Мощность автоматически установлена {required_power} Вт для D{diam}")
            elif cur_power != required_power:
                payload["moshchnost"] = required_power
                messages.append(f"Мощность скорректирована до {required_power} Вт для D{diam}")

    key = build_code_key(payload)
    if not key:
        t_upper = str(payload.get("tip", "")).upper()
        k = payload.get("tip_klapana")
        if k == "pov":
            if not payload.get("raspolozhenie"):
                messages.append("Для поворотного клапана необходимо выбрать расположение: верх/низ")
            if t_upper in ("VBV", "VBA", "VBR"):
                if not payload.get("tip_motora"):
                    messages.append("Для выбранного типа требуется указать тип мотора (6E/6D)")
                if not str(payload.get("moshchnost", "")).strip():
                    messages.append("Для выбранного типа требуется указать мощность (370/750)")
        elif k == "grav":
            if not payload.get("grav_variant"):
                messages.append("Для гравитационного клапана необходимо выбрать вариант: внутренний/внешний")
            if t_upper in ("VBV", "VBA", "VBR"):
                if not payload.get("tip_motora"):
                    messages.append("Для выбранного типа требуется указать тип мотора (6E/6D)")
                if not str(payload.get("moshchnost", "")).strip():
                    messages.append("Для выбранного типа требуется указать мощность (370/750)")
        elif k == "dvustv":
            messages.append("Двустворчатый клапан поддерживается не для всех типов — проверьте выбранный тип шахты")

    result: Dict[str, Dict[str, Any]] = {}
    if key and code_mapping.get(key):
        art = code_mapping[key]
        base = by_art.get(art, {"artikul": art, "name": f"Комплект шахты ({key})"})
        result[art] = {"article": art, "name": base.get("name", ""), "quantity": 1}

    tip_upper = str(tip).upper()
    if not result and tip_upper == "VBP":
        wildcard_key = f"VBP_{diam}_*_*_pov_niz"
        art = code_mapping.get(wildcard_key)
        if art:
            base = by_art.get(art, {"artikul": art, "name": f"Комплект шахты VBP-{diam} (поворотный, низ)"})
            result[art] = {"article": art, "name": base.get("name", ""), "quantity": 1}
        else:
            it = _find(lambda x: _name_has(x, "vbp", diam, "поворот", "низ"))
            if it:
                art = str(it.get("artikul") or it.get("article"))
                result[art] = {"article": art, "name": it.get("name", ""), "quantity": 1}
            else:
                vbp_by_diam = {"560": "89142", "710": "89143", "800": "89153"}
                art = vbp_by_diam.get(diam)
                if art:
                    base = by_art.get(art, {"artikul": art, "name": f"Секция камина VBP-{diam} (2м, поворотный клапан, низ)"})
                    result[art] = {"article": art, "name": base.get("name", ""), "quantity": 1}

    if not result:
        klapan_label = {"pov": "поворотный", "grav": "гравитационный", "dvustv": "двустворчатый"}.get(klapan, klapan or "—")
        messages.append(f"Варианта базовой секции для {tip}-{diam} ({klapan_label}) нет в каталоге/маппинге")

    top = payload.get("verhnyaya_chast")
    if tip_upper in ("VBA", "VBP"):
        if _norm(top) == "rastrub":
            messages.append("Для приточных шахт верхняя часть 'раструб' недоступна — установлен 'зонт'")
            top = "zont"
        elif not _norm(top):
            top = "zont"
            messages.append("Для приточных шахт верхняя часть по умолчанию — 'зонт'")
    elif tip_upper == "VBR":
        if _norm(top) != "zont":
            messages.append("Для VBR верхняя часть всегда 'зонт'")
        top = "zont"

    if top:
        it = pick_top_part(top, diam)
        if it:
            art = str(it.get("artikul") or it.get("article"))
            result[art] = {"article": art, "name": it.get("name", ""), "quantity": 1}
        else:
            if _norm(top) == "rastrub":
                messages.append(f"Раструб для диаметра D{diam} не найден в каталоге")

    if tip_upper == "VBR":
        it = pick_vbr_podmesh(diam)
        if it:
            art = str(it.get("artikul") or it.get("article"))
            result[art] = {"article": art, "name": it.get("name", ""), "quantity": 1}
        else:
            messages.append(f"Секция подмешивания для D{diam} не найдена в каталоге")

    g = payload.get("germetizatsiya") or {}
    if bool(g.get("membrana")):
        it = pick_membrana_by_diam(diam)
        if it:
            art = str(it.get("artikul") or it.get("article"))
            result[art] = {"article": art, "name": it.get("name", ""), "quantity": 1}
        else:
            messages.append(f"Мембрана для диаметра D{diam} не найдена в каталоге")
    if bool(g.get("lenta")):
        it = pick_lenta()
        if it:
            art = str(it.get("artikul") or it.get("article"))
            result[art] = {"article": art, "name": it.get("name", ""), "quantity": 1}
        else:
            messages.append("Лента битумная не найдена в каталоге")

    needs_motor = (tip in ("VBV", "VBA", "VBR")) and (klapan in ("pov", "grav"))
    if bool(payload.get("avtomat")) and needs_motor:
        motor = _norm(payload.get("tip_motora"))
        power = str(payload.get("moshchnost", "")).strip()
        amps_map = {
            ("6e", "370"): 3.0,
            ("6e", "750"): 5.0,
            ("6d", "370"): 1.1,
            ("6d", "750"): 2.4,
        }
        target = amps_map.get((motor, power))
        if target is None:
            messages.append("Не удалось определить номинал автомата: неизвестная комбинация тип мотора × мощность")
        else:
            it = pick_avtomat(target)
            if it:
                art = str(it.get("artikul") or it.get("article"))
                result[art] = {"article": art, "name": it.get("name", ""), "quantity": 1}
            else:
                messages.append(f"Автомат защиты {target} A не найден в каталоге")

    meters = int(payload.get("udlinenie_m") or 0)
    if meters:
        it, qty = pick_udlinenie_sections(diam, meters)
        if it and qty > 0:
            art = str(it.get("artikul") or it.get("article"))
            result[art] = {"article": art, "name": it.get("name", ""), "quantity": qty}

    if bool(payload.get("kapleulavlivatel")):
        if tip_upper in ("VBA", "VBP", "VBR"):
            messages.append("Каплеулавливатель не применяется для приточных шахт и будет проигнорирован")
        else:
            it = pick_kapleu(diam)
            if it:
                art = str(it.get("artikul") or it.get("article"))
                result[art] = {"article": art, "name": it.get("name", ""), "quantity": 1}

    if payload.get("korona") and (payload.get("tip_klapana") != "dvustv"):
        it = pick_korona()
        if it:
            art = str(it.get("artikul") or it.get("article"))
            result[art] = {"article": art, "name": it.get("name", ""), "quantity": 1}

    if bool(payload.get("montazhny_komplekt")):
        it = None
        if klapan == "pov":
            it = by_art.get("89151") or _find(lambda x: _name_has(x, "монтажный", "комплект") and _name_has(x, "vb") and _name_has(x, "поворот"))
        elif klapan == "grav":
            it = by_art.get("89152") or _find(lambda x: _name_has(x, "монтажный", "комплект") and _name_has(x, "vb") and _name_has(x, "гравита"))
        if it:
            art = str(it.get("artikul") or it.get("article"))
            result[art] = {"article": art, "name": it.get("name", ""), "quantity": 1}
        else:
            messages.append("Монтажный комплект для выбранного типа клапана не найден в каталоге")

    if klapan in ("pov", "dvustv"):
        drives = list_available_drives()
        if drives:
            lines = [f"• {str(d.get('artikul') or d.get('article'))} — {d.get('name','')}" for d in drives]
            listing = "&lt;br&gt;".join(lines)
            messages.append(f"Не забудьте добавить привод! Доступные приводы:&lt;br&gt;{listing}")
        else:
            messages.append("Не забудьте добавить привод! (в каталоге приводы не найдены)")

    out = list(result.values())
    return out, messages



@app.route("/api/select", methods=["POST"])
def api_select():
    payload = request.get_json(silent=True) or {}
    results, messages = _select_components(payload)
    if not results:
        base_reason = "; ".join(messages) if messages else "Основание подбора не найдено в каталоге/маппинге"
        return jsonify({"results": [], "message": f"Ничего не найдено. {base_reason}"})
    if messages:
        return jsonify({"results": results, "message": "; ".join(messages)})
    return jsonify({"results": results})


# --- New export route ---
@app.route("/api/export", methods=["POST"])
def api_export():
    payload = request.get_json(silent=True) or {}
    results, messages = _select_components(payload)
    if not results:
        msg = "; ".join(messages) if messages else "Ничего не найдено"
        return jsonify({"error": msg}), 400

    # Поддержка опциональных параметров: group, qty_multiplier
    group_title = str(payload.get("group") or "").strip()  # например: "Коридор"
    qty_multiplier = int(payload.get("qty_multiplier") or 1)
    if qty_multiplier < 1:
        qty_multiplier = 1

    wb = Workbook()
    ws = wb.active
    ws.title = "КП"

    # Заголовки по требуемой форме
    headers = ["н/п", "Наименование", "Цена, руб. с НДС", "Кол-во, шт.", "Сумма, руб. с НДС"]
    ws.append(headers)

    # Стили заголовков
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    bold = Font(bold=True)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left = Alignment(horizontal="left", vertical="center", wrap_text=True)
    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    header_fill = PatternFill("solid", fgColor="F2F2F2")

    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = bold
        cell.alignment = center
        cell.border = border
        cell.fill = header_fill

    # Необязательная строка-группа (например, "Коридор")
    row_idx = 2
    if group_title:
        ws.append([group_title, "", "", "", ""])
        for col in range(1, len(headers) + 1):
            ws.cell(row=row_idx, column=col).border = border
            ws.cell(row=row_idx, column=col).font = bold
            ws.cell(row=row_idx, column=1).alignment = left
        row_idx += 1

    # Данные
    total = 0.0
    counter = 1
    for row in results:
        name = row.get("name", "")
        base_qty = float(row.get("quantity") or 0)
        qty = base_qty * qty_multiplier
        art = str(row.get("article") or "").strip()
        # Цена
        price = _price_of(art)
        summa = price * qty
        total += summa

        ws.append([counter, name, price, qty, summa])
        for col in range(1, len(headers) + 1):
            ws.cell(row=row_idx, column=col).border = border
            ws.cell(row=row_idx, column=2).alignment = left
            if col in (1, 3, 4, 5):
                ws.cell(row=row_idx, column=col).alignment = center
        # Числовые форматы (RU): пробел как разделитель тысяч, запятая — десятичная
        ws.cell(row=row_idx, column=3).number_format = '# ##0'
        ws.cell(row=row_idx, column=4).number_format = '# ##0'
        ws.cell(row=row_idx, column=5).number_format = '# ##0'
        row_idx += 1
        counter += 1

    # ИТОГО
    ws.append(["", "", "", "ИТОГО, руб. с НДС", total])
    for col in range(1, len(headers) + 1):
        ws.cell(row=row_idx, column=col).border = border
        if col in (4, 5):
            ws.cell(row=row_idx, column=col).font = bold
            ws.cell(row=row_idx, column=col).alignment = center
        if col == 5:
            ws.cell(row=row_idx, column=col).number_format = '# ##0'
    row_idx += 1

    # Ширины колонок и выравнивание
    ws.column_dimensions['A'].width = 6   # н/п
    ws.column_dimensions['B'].width = 70  # Наименование
    ws.column_dimensions['C'].width = 18  # Цена
    ws.column_dimensions['D'].width = 12  # Кол-во
    ws.column_dimensions['E'].width = 20  # Сумма

    if messages:
        ws2 = wb.create_sheet("Комментарии")
        from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
        bold = Font(bold=True)
        center = Alignment(horizontal="center", vertical="center", wrap_text=True)
        left = Alignment(horizontal="left", vertical="center", wrap_text=True)
        thin = Side(style="thin")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        header_fill = PatternFill("solid", fgColor="F2F2F2")

        # Заголовки как у основного листа
        headers2 = ["н/п", "Наименование", "Цена, руб. с НДС", "Кол-во, шт.", "Сумма, руб. с НДС"]
        ws2.append(headers2)
        for col in range(1, len(headers2) + 1):
            cell = ws2.cell(row=1, column=col)
            cell.font = bold
            cell.alignment = center
            cell.border = border
            cell.fill = header_fill

        row_c = 2
        counter_c = 1

        # Если среди сообщений есть общий заголовок (например, про приводы) — выведем строкой-группой
        title_written = False
        for msg in messages:
            if "Не забудьте добавить привод" in str(msg):
                ws2.append(["Не забудьте добавить привод!", "", "", "", ""])
                for col in range(1, len(headers2) + 1):
                    ws2.cell(row=row_c, column=col).border = border
                    ws2.cell(row=row_c, column=col).font = bold
                    ws2.cell(row=row_c, column=1).alignment = left
                row_c += 1
                title_written = True
                break

        import re
        leftovers: list[str] = []

        # Разбираем каждое сообщение, ищем маркеры вида "• 7547 — Электропривод ..."
        for msg in messages:
            text = str(msg).replace("&lt;br&gt;", "\n").replace("<br>", "\n")
            # Если это список с точками — распарсим построчно
            lines = [ln.strip() for ln in text.splitlines() if ln.strip()]
            parsed_any = False
            for ln in lines:
                # Уберём лидирующие маркеры
                ln_clean = ln.lstrip("•-— ").strip()
                m = re.match(r"^(\d+)\s*[—-]\s*(.+)$", ln_clean)
                if not m:
                    continue
                art = m.group(1).strip()
                name = m.group(2).strip()
                price = _price_of(art)
                qty = 1.0
                summa = price * qty

                ws2.append([counter_c, f"{art} — {name}", price, qty, summa])
                for col in range(1, len(headers2) + 1):
                    ws2.cell(row=row_c, column=col).border = border
                    ws2.cell(row=row_c, column=2).alignment = left
                    if col in (1, 3, 4, 5):
                        ws2.cell(row=row_c, column=col).alignment = center
                ws2.cell(row=row_c, column=3).number_format = '# ##0'
                ws2.cell(row=row_c, column=4).number_format = '# ##0'
                ws2.cell(row=row_c, column=5).number_format = '# ##0'
                row_c += 1
                counter_c += 1
                parsed_any = True

            if not parsed_any and text:
                leftovers.append(text)

        # Если есть непреобразованные сообщения — добавим их отдельным блоком ниже таблицы
        if leftovers:
            # Пустая строка-разделитель
            ws2.append(["", "", "", "", ""])
            for col in range(1, len(headers2) + 1):
                ws2.cell(row=row_c, column=col).border = border
            row_c += 1

            ws2.append(["Примечания", "", "", "", ""])
            for col in range(1, len(headers2) + 1):
                ws2.cell(row=row_c, column=col).border = border
                ws2.cell(row=row_c, column=col).font = bold
            row_c += 1

            for text in leftovers:
                # Каждое примечание — одной ячейкой в колонке B, во всех остальных — прочерки
                ws2.append(["", text, "", "", ""])
                ws2.cell(row=row_c, column=2).alignment = left
                for col in range(1, len(headers2) + 1):
                    ws2.cell(row=row_c, column=col).border = border
                row_c += 1

        # Ширины колонок как у основного листа
        ws2.column_dimensions['A'].width = 6
        ws2.column_dimensions['B'].width = 70
        ws2.column_dimensions['C'].width = 18
        ws2.column_dimensions['D'].width = 12
        ws2.column_dimensions['E'].width = 20

    stream = BytesIO()
    wb.save(stream)
    stream.seek(0)
    filename = f"KP_{payload.get('tip','X')}_{payload.get('diametr','D')}.xlsx"
    return send_file(stream, as_attachment=True, download_name=filename, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


@app.route("/", methods=["GET"])
def index():
    return send_from_directory("web", "index.html")


@app.route("/web/<path:filename>")
def static_files(filename):
    return send_from_directory("web", filename)


@app.route("/favicon.ico")
def favicon():
    return send_from_directory("web", "agroventlabel.png", mimetype='image/png')


if __name__ == "__main__":
    # PROD: запускать через gunicorn/uwsgi; debug только локально
    app.run(host="0.0.0.0", port=5001, debug=True)